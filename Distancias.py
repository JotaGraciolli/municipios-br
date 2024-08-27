import googlemaps
import openpyxl
from geopy.geocoders import Nominatim

# Insira sua chave de API do Google Maps
gmaps = googlemaps.Client(key='AIzaSyCboUFVqK8j-iUZoHsYXedLR9ZdvH3DvtI')

# Preparando o objeto ws (Worksheet) para o looping
FilePath = r'C:\Users\sidnei.graciolli\OneDrive - Avanade\Documents\PythonProjects\Distancias\Cidades.xlsx'
DataFile = openpyxl.load_workbook(FilePath)
ws = DataFile['Dist']

# Objeto que auxilia a determinar a distância nos casos em que a API do Google Maps retorna 'NOT_FOUND'
# Esses casos acontecem quando a API do Google não consegue geolocalizar pelo nome da cidade
geolocator = Nominatim(user_agent="CalcDist")

# Looping pela planilha
for RowNum, RowVal in enumerate(ws.iter_rows(max_col=8,min_row=2,max_row=ws.max_row)):

    # Chamada da API utilizando a origem e o destino
    origem = ws.cell(row=RowNum+2, column=1).value
    destino = ws.cell(row=RowNum+2, column=2).value
    resultado = gmaps.distance_matrix(origem, destino, mode='driving')
    
    # Tratamento dos possíveis retornos
    if resultado['rows'][0]['elements'][0]['status'] == 'OK':
        ws.cell(row=RowNum+2, column=3).value = resultado['rows'][0]['elements'][0]['distance']['text']
    elif resultado['rows'][0]['elements'][0]['status'] == 'ZERO_RESULTS':
        ws.cell(row=RowNum+2, column=3).value = 'Nenhuma rota terrestre encontrada'
    else:
        locationOrigem = geolocator.geocode(origem)
        resultado = gmaps.distance_matrix(locationOrigem, destino, mode='driving')
        ws.cell(row=RowNum+2, column=3).value = resultado['rows'][1]['elements'][0]['distance']['text']

# Salva a planiha
DataFile.save(FilePath)