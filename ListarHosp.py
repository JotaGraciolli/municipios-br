# %%
import requests
import openpyxl
import time
from tqdm import tqdm 

# %%
# Declaração de variáveis

linhaInicial = 3
linha = 2
countMunicipio = 1
SleepTime = 1
FilePath = r'C:\Users\sidnei.graciolli\OneDrive - Avanade\WORK\git\MunicipiosBr\MunicipiosBr.xlsx'
api_key = 'AIzaSyCboUFVqK8j-iUZoHsYXedLR9ZdvH3DvtI'
url = "https://maps.googleapis.com/maps/api/place/textsearch/json"

# %%
def Populate(pUrl, pParams, pWsRes, pWs, pRowNum, pObs):
    global linha

    response = requests.get(pUrl, params=pParams)
    results = response.json().get('results', [])

    for result in results:
        pWsRes.cell(row=linha, column=1).value = pWs.cell(row=pRowNum+linhaInicial, column=7).value
        pWsRes.cell(row=linha, column=2).value = pWs.cell(row=pRowNum+linhaInicial, column=2).value
        pWsRes.cell(row=linha, column=3).value = result.get('name')
        pWsRes.cell(row=linha, column=4).value = result.get('formatted_address')
        pWsRes.cell(row=linha, column=5).value = pObs
        
        linha += 1

    return response.json().get('next_page_token')

# %%
DataFile = openpyxl.load_workbook(FilePath)
ws = DataFile['BaseMunicipios']

wsRes = DataFile['NosocomiosPorMunicipio']

totalMunicipio = ws.max_row

for RowNum, RowVal in tqdm(enumerate(ws.iter_rows(max_col=15,min_row=linhaInicial,max_row=totalMunicipio)), 
                           desc='Processando...', total=totalMunicipio-linhaInicial):
    municipio = f'{ws.cell(row=RowNum+linhaInicial, column=7).value}, {ws.cell(row=RowNum+linhaInicial, column=1).value}, Brasil'
    strQuery = f'hospitais em {municipio}'

    ws.cell(row=RowNum+linhaInicial, column=11).value = strQuery

    # Parâmetros da requisição
    params = {
        'query': strQuery,
        'key': api_key
    }

    next_page_token = Populate(url, params, wsRes, ws, RowNum, strQuery)

    #time.sleep(SleepTime)

    while next_page_token:
        time.sleep(SleepTime)
        params['pagetoken'] = next_page_token
        next_page_token = Populate(url, params, wsRes, ws, RowNum, strQuery)

DataFile.save(FilePath)